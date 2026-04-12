from django.db.models import Sum
from django.shortcuts import render
from django.core.cache import cache
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from .models import Employee, CashAdvance, AdvanceReport
from .forms import SubreportForm
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Импорт для PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm


def subreport_view(request):
    """Основное представление отчета по подотчетникам"""
    
    # Восстанавливаем параметры из сессии при GET без параметров
    if request.method == 'GET' and not request.GET:
        saved_params = request.session.get('subreport_params', {})
        form = SubreportForm(saved_params or {'date': datetime.date.today().isoformat()})
    else:
        form = SubreportForm(request.GET)
        if form.is_valid():
            # Сохраняем параметры в сессию (исключая export)
            params_to_save = {k: v for k, v in request.GET.dict().items() if k != 'export'}
            request.session['subreport_params'] = params_to_save
    
    report_data = []
    total_given = 0
    total_reported = 0
    total_debt = 0
    report_date = None
    
    if form.is_valid():
        report_date = form.cleaned_data['date']
        department = form.cleaned_data.get('department', '').strip()
        hide_zero_debt = form.cleaned_data.get('hide_zero_debt', False)
        sort_by = form.cleaned_data.get('sort_by', 'debt_desc')
        show_only_debtors = form.cleaned_data.get('show_only_debtors', False)
        
        # Создаем ключ кэша
        cache_key = f"subreport_{report_date}_{department}_{hide_zero_debt}_{sort_by}_{show_only_debtors}"
        report_data = cache.get(cache_key)
        
        # Если данных нет в кэше, формируем отчет
        if report_data is None:
            report_data = []
            
            # Оптимизированный запрос сотрудников
            employees = Employee.objects.all()
            
            # Фильтрация по отделу
            if department:
                employees = employees.filter(department__icontains=department)
            
            # Предзагрузка всех выдач и отчетов для оптимизации
            cash_advances = CashAdvance.objects.filter(date__lte=report_date).select_related('employee')
            advance_reports = AdvanceReport.objects.filter(date__lte=report_date).select_related('employee')
            
            # Группировка по сотрудникам
            given_dict = {}
            reported_dict = {}
            
            for ca in cash_advances:
                given_dict[ca.employee_id] = given_dict.get(ca.employee_id, 0) + float(ca.amount)
            
            for ar in advance_reports:
                reported_amount = float(ar.expense_amount or 0) + float(ar.return_amount or 0) if hasattr(ar, 'return_amount') else float(ar.expense_amount or 0)
                reported_dict[ar.employee_id] = reported_dict.get(ar.employee_id, 0) + reported_amount
            
            # Расчет данных для каждого сотрудника
            for emp in employees:
                given = given_dict.get(emp.id, 0)
                reported = reported_dict.get(emp.id, 0)
                debt = given - reported
                
                # Пропускаем нулевую задолженность если нужно
                if hide_zero_debt and abs(debt) < 0.01:
                    continue
                
                # Показываем только должников
                if show_only_debtors and debt <= 0:
                    continue
                
                report_data.append({
                    'employee': emp,
                    'given': round(given, 2),
                    'reported': round(reported, 2),
                    'debt': round(debt, 2),
                })
            
            # Сортировка данных
            if sort_by == 'debt_desc':
                report_data.sort(key=lambda x: x['debt'], reverse=True)
            elif sort_by == 'debt_asc':
                report_data.sort(key=lambda x: x['debt'])
            elif sort_by == 'name':
                report_data.sort(key=lambda x: x['employee'].full_name())
            
            # Сохраняем в кэш на 15 минут
            cache.set(cache_key, report_data, 60 * 15)
        
        # Подсчет итоговых сумм
        for row in report_data:
            total_given += row['given']
            total_reported += row['reported']
            total_debt += row['debt']
        
        total_given = round(total_given, 2)
        total_reported = round(total_reported, 2)
        total_debt = round(total_debt, 2)
    
    # Обработка экспорта
    export_format = request.GET.get('export')
    
    if export_format == 'excel' and report_data and form.is_valid():
        return export_to_excel(report_data, total_given, total_reported, total_debt, report_date)
    
    if export_format == 'pdf' and report_data and form.is_valid():
        return export_to_pdf_reportlab(report_data, total_given, total_reported, total_debt, report_date)
    
    # Подготовка контекста для шаблона
    context = {
        'form': form,
        'report_data': report_data,
        'total_given': total_given,
        'total_reported': total_reported,
        'total_debt': total_debt,
        'report_date': report_date,
        'has_data': len(report_data) > 0,
    }
    
    return render(request, 'repo/subreport.html', context)


def export_to_excel(report_data, total_given, total_reported, total_debt, report_date):
    """Экспорт отчета в Excel с профессиональным оформлением"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Подотчетники"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    debtor_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки таблицы
    headers = ['№', 'Сотрудник', 'Отдел', 'Должность', 'Получено под отчет, ₽', 
               'Отчитано/Возвращено, ₽', 'Задолженность, ₽']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполнение данных
    for i, row in enumerate(report_data, 1):
        # Номер
        cell = ws.cell(row=i+1, column=1, value=i)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # ФИО
        cell = ws.cell(row=i+1, column=2, value=row['employee'].full_name())
        cell.border = border
        
        # Отдел
        cell = ws.cell(row=i+1, column=3, value=row['employee'].department or '')
        cell.border = border
        
        # Должность
        cell = ws.cell(row=i+1, column=4, value=row['employee'].position or '')
        cell.border = border
        
        # Получено
        cell = ws.cell(row=i+1, column=5, value=row['given'])
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Отчитано
        cell = ws.cell(row=i+1, column=6, value=row['reported'])
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Задолженность
        cell = ws.cell(row=i+1, column=7, value=row['debt'])
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right')
        cell.border = border
        
        # Выделение должников желтым цветом
        if row['debt'] > 0:
            for col in range(1, 8):
                ws.cell(row=i+1, column=col).fill = debtor_fill
    
    # Итоговая строка
    last_row = len(report_data) + 2
    ws.cell(row=last_row, column=4, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=float(total_given)).font = Font(bold=True)
    ws.cell(row=last_row, column=5).number_format = '#,##0.00'
    ws.cell(row=last_row, column=6, value=float(total_reported)).font = Font(bold=True)
    ws.cell(row=last_row, column=6).number_format = '#,##0.00'
    ws.cell(row=last_row, column=7, value=float(total_debt)).font = Font(bold=True)
    ws.cell(row=last_row, column=7).number_format = '#,##0.00'
    
    for col in range(4, 8):
        ws.cell(row=last_row, column=col).fill = total_fill
        ws.cell(row=last_row, column=col).border = border
    
    # Настройка ширины колонок
    column_widths = [5, 30, 20, 25, 18, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Формирование ответа
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="subreport_{report_date}.xlsx"'
    wb.save(response)
    return response


def export_to_pdf_reportlab(report_data, total_given, total_reported, total_debt, report_date):
    """Экспорт в PDF с помощью reportlab (не требует системных зависимостей)"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="subreport_{report_date}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), 
                           rightMargin=10*mm, leftMargin=10*mm,
                           topMargin=15*mm, bottomMargin=15*mm)
    
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']
    
    story = []
    
    # Заголовок
    story.append(Paragraph("Отчет по подотчетным лицам", title_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"Дата отчета: {report_date.strftime('%d.%m.%Y')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Данные для таблицы
    table_data = [['№', 'Сотрудник', 'Должность', 'Отдел', 'Получено, ₽', 'Отчитано, ₽', 'Задолженность, ₽']]
    
    for i, row in enumerate(report_data, 1):
        table_data.append([
            str(i),
            row['employee'].full_name(),
            row['employee'].position or '—',
            row['employee'].department or '—',
            f"{row['given']:.2f}",
            f"{row['reported']:.2f}",
            f"{row['debt']:.2f}"
        ])
    
    # Итоговая строка
    table_data.append([
        '', '', '', 'ИТОГО:',
        f"{total_given:.2f}",
        f"{total_reported:.2f}",
        f"{total_debt:.2f}"
    ])
    
    # Создание таблицы
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.green),
    ]))
    
    # Выделение должников желтым
    for i in range(1, len(table_data) - 1):
        try:
            if float(table_data[i][6]) > 0:
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF3CD')),
                ]))
        except (ValueError, IndexError):
            pass
    
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Сформировано автоматически: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    
    doc.build(story)
    return response